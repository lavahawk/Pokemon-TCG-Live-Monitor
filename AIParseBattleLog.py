import json
import os
import re
import sys
import time
import tkinter as tk
import warnings
from tkinter import simpledialog

import openpyxl
import pyperclip
from pydantic import BaseModel

from BattleDatabase import BattleDatabase
from app_settings import (
    is_local_only_mode,
    load_api_key,
    load_username,
    save_username,
)


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "TCGExampleSheet.xlsx")
DECK_SHEET = "Limitless Meta"
LIMITLESS_STANDARD_CACHE_FILE = os.path.join(BASE_DIR, ".meta_cache_limitless_standard.json")
LAST_DECK_FILE = os.path.join(BASE_DIR, ".last_deck")
LAST_RANK_FILE = os.path.join(BASE_DIR, ".last_rank")
LOG_DIR = os.path.join(BASE_DIR, "Logs")
# Shared flag file written by the Limitless dashboard manager (StatsUI) when
# the player is actively in a tournament match. Contains JSON:
#   {"in_tournament": true, "opponent": "<opponent username>"}
TOURNAMENT_FLAG_FILE = os.path.join(BASE_DIR, ".in_tournament")

db = BattleDatabase()


def load_tournament_state():
    """Read the shared tournament flag file written by the Limitless manager.

    Returns a dict {"in_tournament": bool, "opponent": str} or an empty dict
    if no tournament match is currently active.
    """
    if not os.path.exists(TOURNAMENT_FLAG_FILE):
        return {}
    try:
        with open(TOURNAMENT_FLAG_FILE, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
        if isinstance(payload, dict) and payload.get("in_tournament"):
            return {
                "in_tournament": True,
                "opponent": str(payload.get("opponent") or "").strip(),
            }
    except Exception:
        pass
    return {}


def parse_opponent_from_log(battlelog, username):
    """Extract the opponent's username from a battle log.

    The battle log contains both players' usernames (e.g. "PlayerA chose
    tails...", "PlayerB won the coin toss."). The opponent is the username
    that is NOT the player's own username.
    """
    if not battlelog or not username:
        return None
    own = normalize_name(username)
    # Collect all usernames that appear as actors in the log.
    candidates = set()
    # "X chose heads/tails for the opening coin flip."
    for match in re.findall(r"([A-Za-z0-9_\- ]+?)\s+chose\s+(?:heads|tails)\s+for the opening coin flip", battlelog, flags=re.IGNORECASE):
        candidates.add(match.strip())
    # "X won the coin toss."
    for match in re.findall(r"([A-Za-z0-9_\- ]+?)\s+won the coin toss", battlelog, flags=re.IGNORECASE):
        candidates.add(match.strip())
    # "X decided to go first/second."
    for match in re.findall(r"([A-Za-z0-9_\- ]+?)\s+decided to go (?:first|second)", battlelog, flags=re.IGNORECASE):
        candidates.add(match.strip())
    # "X's Turn"
    for match in re.findall(r"([A-Za-z0-9_\- ]+?)'s\s+Turn", battlelog, flags=re.IGNORECASE):
        candidates.add(match.strip())
    # "X wins."
    for match in re.findall(r"([A-Za-z0-9_\- ]+?)\s+wins\.", battlelog, flags=re.IGNORECASE):
        candidates.add(match.strip())

    for candidate in candidates:
        if candidate and normalize_name(candidate) != own:
            return candidate
    return None


def is_tournament_battle(battlelog, username):
    """Determine whether a battle is a Limitless tournament game.

    A battle is only tagged as a tournament game if the player is actively in
    a tournament match (per the shared flag file) AND the battle-log opponent
    matches the tournament opponent. This prevents false positives from simply
    being enrolled in a tournament days in advance.
    """
    state = load_tournament_state()
    if not state.get("in_tournament"):
        return False
    tournament_opponent = state.get("opponent")
    if not tournament_opponent:
        # No opponent known — be conservative and don't tag.
        return False
    log_opponent = parse_opponent_from_log(battlelog, username)
    if not log_opponent:
        return False
    return normalize_name(log_opponent) == normalize_name(tournament_opponent)


def prompt_string(title, prompt, *, secret=False):
    if sys.stdin.isatty():
        return input(f"{prompt}\n> ").strip()
    root = tk.Tk()
    root.withdraw()
    try:
        return simpledialog.askstring(title, prompt, show="*" if secret else None)
    finally:
        root.destroy()


def ensure_username():
    username = load_username()
    if username:
        return username
    username = prompt_string(
        "Enter Your TCG Live Username",
        "No username was found.\n\nPlease enter your TCG Live username:",
    )
    if username:
        save_username(username.strip())
    return (username or "").strip()


def load_possible_decks():
    cache_rows = []
    if os.path.exists(LIMITLESS_STANDARD_CACHE_FILE):
        try:
            with open(LIMITLESS_STANDARD_CACHE_FILE, "r", encoding="utf-8") as handle:
                payload = json.load(handle)
            rows = payload.get("rows", []) if isinstance(payload, dict) else []
            if isinstance(rows, list):
                cache_rows = rows
        except Exception:
            cache_rows = []

    if cache_rows:
        ranked = sorted(
            cache_rows,
            key=lambda row: int((row or {}).get("count", 0) or 0),
            reverse=True,
        )
        names = []
        seen = set()
        for row in ranked:
            deck_name = str((row or {}).get("deck") or "").strip()
            if not deck_name or deck_name.lower() == "other":
                continue
            normalized = normalize_name(deck_name)
            if normalized in seen:
                continue
            seen.add(normalized)
            names.append(deck_name)
            if len(names) >= 50:
                break
        if names:
            return names

    if not os.path.exists(EXCEL_FILE):
        return []
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                category=DeprecationWarning,
                message="Pyarrow will become a required dependency of pandas.*",
            )
            import pandas as pd
        deck_df = pd.read_excel(EXCEL_FILE, DECK_SHEET, usecols="B", skiprows=0, nrows=60)
        return [str(value).strip() for value in deck_df.iloc[:, 0].dropna().tolist() if str(value).strip()]
    except Exception:
        return []


def normalize_name(value):
    return re.sub(r"\s+", " ", (value or "").strip()).casefold()


def parse_winner_from_log(battlelog):
    if not battlelog:
        return None
    explicit = re.findall(r"All Prize cards taken\.\s*(.+?) wins\.", battlelog, flags=re.IGNORECASE)
    if explicit:
        return explicit[-1].strip()
    generic = re.findall(r"\b(.+?) wins\.", battlelog, flags=re.IGNORECASE)
    return generic[-1].strip() if generic else None


def parse_result_from_log(battlelog, username):
    winner = parse_winner_from_log(battlelog)
    if not winner or not username:
        return None
    return "Win" if normalize_name(winner) == normalize_name(username) else "Loss"


def load_last_deck():
    if not os.path.exists(LAST_DECK_FILE):
        return None
    try:
        with open(LAST_DECK_FILE, "r", encoding="utf-8") as handle:
            value = handle.read().strip()
        return value or None
    except Exception:
        return None


def load_last_rank():
    if not os.path.exists(LAST_RANK_FILE):
        return None
    try:
        with open(LAST_RANK_FILE, "r", encoding="utf-8") as handle:
            return int(handle.read().strip())
    except Exception:
        return None


def latest_log_file():
    if not os.path.isdir(LOG_DIR):
        return None
    log_files = [f for f in os.listdir(LOG_DIR) if f.startswith("battle_log_") and f.endswith(".txt")]
    if not log_files:
        return None
    log_files.sort(reverse=True)
    return os.path.join(LOG_DIR, log_files[0])


def read_battlelog_file(file_path):
    if not file_path:
        return None
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as handle:
            return handle.read().strip()
    except Exception:
        return None


def save_to_excel(my_deck, opponents_deck, win_or_loss):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet_name = (my_deck or "Unknown Deck")[:31]
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(title=sheet_name)

    next_row = 1
    while sheet.cell(row=next_row, column=1).value is not None:
        next_row += 1

    sheet.cell(row=next_row, column=1).value = win_or_loss
    sheet.cell(row=next_row, column=2).value = opponents_deck
    workbook.save(EXCEL_FILE)


def save_to_database(my_deck, opponents_deck, win_or_loss, confidence, log_file_path=None, is_tournament=False):
    db.add_battle(
        my_deck=my_deck,
        opponent_deck=opponents_deck,
        result=win_or_loss,
        my_rank=load_last_rank(),
        confidence=confidence,
        deck_source="AI" if not is_local_only_mode() and load_api_key() else "Local",
        log_file=os.path.abspath(log_file_path) if log_file_path else latest_log_file(),
        is_tournament=is_tournament,
    )


def local_fallback_parse(battlelog, username):
    my_deck = load_last_deck()
    if not my_deck:
        my_deck = prompt_string(
            "Enter Your Deck",
            "Local-only mode could not find your deck from OCR.\n\nPlease enter your deck name:",
        )

    opponents_deck = prompt_string(
        "Enter Opponent Deck",
        "Local-only mode is active.\n\nPlease enter the opponent's deck name for this battle:",
    )

    win_or_loss = parse_result_from_log(battlelog, username)
    if not win_or_loss:
        win_or_loss = prompt_string(
            "Enter Result",
            "The battle result could not be inferred from the log.\n\nEnter Win or Loss:",
        )

    confidence = 90
    if not load_last_deck():
        confidence = 80
    if not parse_result_from_log(battlelog, username):
        confidence = min(confidence, 75)

    return {
        "My_deck": (my_deck or "").strip(),
        "OpponentsDeck": (opponents_deck or "").strip(),
        "Win_or_Loss": (win_or_loss or "").strip().title(),
        "Confidence": confidence,
    }


class BattleLogOutput(BaseModel):
    My_deck: str
    OpponentsDeck: str
    Win_or_Loss: str
    Confidence: int


def _flatten_scalar(value):
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, list):
        for item in value:
            flattened = _flatten_scalar(item)
            if flattened not in (None, ""):
                return flattened
        return None
    if isinstance(value, dict):
        for key in ("value", "score", "confidence", "percent", "percentage", "raw", "text"):
            if key in value:
                flattened = _flatten_scalar(value.get(key))
                if flattened not in (None, ""):
                    return flattened
        for nested in value.values():
            flattened = _flatten_scalar(nested)
            if flattened not in (None, ""):
                return flattened
        return None
    return str(value).strip()


def _coerce_text_field(payload, *keys):
    for key in keys:
        if key not in payload:
            continue
        flattened = _flatten_scalar(payload.get(key))
        if flattened is None:
            continue
        text = str(flattened).strip()
        if text:
            return text
    return ""


def _coerce_confidence_value(value, default=0):
    flattened = _flatten_scalar(value)
    if flattened is None or flattened == "":
        return int(default)
    if isinstance(flattened, bool):
        return 100 if flattened else int(default)
    if isinstance(flattened, (int, float)):
        numeric = float(flattened)
    else:
        match = re.search(r"-?\d+(?:\.\d+)?", str(flattened))
        if not match:
            return int(default)
        numeric = float(match.group(0))
    return max(0, min(100, int(round(numeric))))


def ai_parse(battlelog, username, api_key):
    from openai import OpenAI

    client = OpenAI(api_key=api_key)
    possible_decks = load_possible_decks()
    prompt = f"""
Here is a list of the current top Pokemon TCG deck names from our latest Limitless top-50-by-count scrape. Prefer matching to one of these when the log supports it: {possible_decks}.
The main attacker of each deck is usually the deck name but not always. Sometimes a deck is named for its engine rather than attacker.
Below is a battle log. Determine the winner and the decks used by each player based on the cards used and held.
If no battlelog was provided return blank.
Please include a confidence value 0-100 that you got each deck correctly.
My username is {username}. Report if I won or lost the battle.
Export the data as JSON with these fields: My_deck, OpponentsDeck, Win_or_Loss, Confidence
Battlelog:
\"\"\"{battlelog}\"\"\"
"""

    completion = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": "You parse Pokemon TCG battle logs and return JSON only.",
            },
            {"role": "user", "content": prompt},
        ],
        response_format={"type": "json_object"},
    )
    payload = json.loads(completion.choices[0].message.content)
    parsed = BattleLogOutput(
        My_deck=_coerce_text_field(payload, "My_deck", "MyDeck", "my_deck"),
        OpponentsDeck=_coerce_text_field(payload, "OpponentsDeck", "OpponentDeck", "opponent_deck", "Opponent"),
        Win_or_Loss=_coerce_text_field(payload, "Win_or_Loss", "WinLoss", "Result", "result").title(),
        Confidence=_coerce_confidence_value(payload.get("Confidence")),
    )
    return {
        "My_deck": parsed.My_deck,
        "OpponentsDeck": parsed.OpponentsDeck,
        "Win_or_Loss": parsed.Win_or_Loss,
        "Confidence": parsed.Confidence,
    }


def main():
    username = ensure_username()
    explicit_log_path = os.path.abspath(sys.argv[1]) if len(sys.argv) > 1 else None
    battlelog = read_battlelog_file(explicit_log_path)
    if not battlelog:
        battlelog = pyperclip.paste()
    if not battlelog and not explicit_log_path:
        latest_path = latest_log_file()
        battlelog = read_battlelog_file(latest_path)
        explicit_log_path = latest_path
    if not battlelog:
        print("No battle log found in clipboard.")
        time.sleep(2)
        return

    settings_local_only = is_local_only_mode()
    api_key = load_api_key()
    use_local_only = settings_local_only or not api_key

    if use_local_only:
        if settings_local_only:
            print("Local-only mode is enabled. Falling back to manual opponent deck entry.")
        else:
            print("No OpenAI API key found. Falling back to local-only parsing.")
        response_dict = local_fallback_parse(battlelog, username)
    else:
        try:
            response_dict = ai_parse(battlelog, username, api_key)
        except Exception as exc:
            print(f"AI parse failed: {exc}")
            print("Falling back to local-only parsing.")
            response_dict = local_fallback_parse(battlelog, username)

    my_deck = response_dict.get("My_deck") or "Unknown Deck"
    opponents_deck = response_dict.get("OpponentsDeck") or "Unknown"
    win_or_loss = (response_dict.get("Win_or_Loss") or "Loss").title()
    confidence = int(response_dict.get("Confidence") or 0)

    print("--------------------------")
    print(
        f"My Deck: {my_deck}  |  Opponent's Deck: {opponents_deck}  |  "
        f"Result: {win_or_loss}  |  Confidence: {confidence}%"
    )
    print("--------------------------")

    try:
        if os.path.exists(EXCEL_FILE):
            save_to_excel(my_deck, opponents_deck, win_or_loss)
            print("Saved successfully to workbook.")
    except Exception as exc:
        print(f"Warning: Could not update workbook: {exc}")

    try:
        save_to_database(my_deck, opponents_deck, win_or_loss, confidence, explicit_log_path)
        print("Battle saved to database.")
    except Exception as exc:
        print(f"Warning: Could not save to database: {exc}")

    time.sleep(3)


if __name__ == "__main__":
    main()
